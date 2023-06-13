import pandas as pd
import numpy as np
from catboost import CatBoostRegressor
from openpyxl import load_workbook
from pathlib import Path
import itertools as it

# Проверка наличия загруженной модели
if 'model' not in globals():
    model = CatBoostRegressor()
    model.load_model('./model_params.cbm')

def top5(name):
    lamb = 0.8
    alp = 1

    def to_flt(s):
        if type(s) != type('help'):
            return s
        s = s.replace(' ', '')
        s = s.replace('\xa0', '')
        s = s.split(',')
        s = '.'.join(s)
        return float(s)

    def prob_counter(alp, lamb, i, j):  # i - номер района, j - номер магазина (порядковый, не ID !!!)
        sum = 0
        for v in range(cntstr):
            distutil = to_flt((dists[('Store' + str(main_data['ID'][v]))][i - 1]))
            sum += to_flt(main_data['Square'][v]) ** alp / (distutil if distutil < 25 else distutil / 5) ** lamb
        distutil = to_flt((dists[('Store' + str(main_data['ID'][j - 1]))][i - 1]))
        return (dists['Population'][i - 1]) * to_flt(main_data['Square'][j - 1]) ** alp / (
            distutil if distutil < 25 else distutil / 5) ** lamb / sum

    def nearest_neighbour(lon, lat):
        neig_id = 0
        mindist = 100000
        for i in range(cntstr):
            dist = abs(lon - to_flt(id_to_address[main_data['ID'][i]][0])) + abs(
                lat - to_flt(id_to_address[main_data['ID'][i]][1]))
            if (dist < mindist):
                mindist = dist
                neig_id = i
        return neig_id

    def data_huff_pred(i, lamb, aph):
        sum = 0
        ind = nearest_neighbour(dataset['Latitude'][i], dataset['Longitude'][i])
        for j in range(112):
            znam = 0
            for v in range(cntstr):
                distutil = to_flt((dists[('Store' + str(main_data['ID'][v]))][j]))

                znam += to_flt(main_data['Square'][v]) ** alp / (distutil if distutil < 25 else distutil / 5) ** lamb

            distutil = to_flt((dists[('Store' + str(main_data['ID'][ind]))][j]))
            sum += (dists['Population'][j]) * to_flt(dataset['Square'][i]) ** alp / (
                distutil if distutil < 25 else distutil / 5) ** lamb / znam
        return sum

    id_to_address = [[0, 0]]
    ws = load_workbook(Path('market-coordinates.xlsx'))['Sheet1']
    address_to_id = {}
    for i, row in it.islice(enumerate(ws.rows), 1, 1000):
        id_to_address.append([row[3].value, row[2].value])
        address_to_id[row[1].value] = int(row[0].value)

    ws = load_workbook(Path('filled-table.xlsx'), read_only=True)['Sheet']

    market_address_ids = []
    market_visitor_frequencies = []

    for i, row in it.islice(enumerate(ws.rows), 229):
        address = ' '.join([str(row[i].value) for i in [4, 3, 2]])
        market_address_ids.append([row[0].value, address_to_id[address]])

        visitors = int(row[14].value)
        try:
            days = int(row[15].value)
        except TypeError:
            days = 30
        market_visitor_frequencies.append(visitors / days)

    market_address_ids = np.array(market_address_ids)
    market_visitor_frequencies = np.array(market_visitor_frequencies)

    main_data = pd.read_csv('./main_data_10.csv')

    cntstr = main_data.shape[0]

    id = []
    for i in range(cntstr):
        for j in range(229):
            if (main_data['Num'][i] == market_address_ids[j][0]):
                id.append(market_address_ids[j][1])
                break
            j += 1
        i += 1

    main_data.insert(1, 'ID', id)

    dists = pd.read_csv('./fixed_dist.csv')

    cat_cols = ['Type', 'Name', 'Building']
    num_cols = ['Square', 'Huff_predict']
    target_col = 'FPD'

    dataset = pd.read_csv('./for_economists.csv')
    data_to_model = pd.read_csv('./for_economists.csv')

    data_to_model = data_to_model.drop('Id', axis=1)
    data_to_model = data_to_model.drop('Price', axis=1)
    data_to_model = data_to_model.drop('Address', axis=1)
    data_to_model = data_to_model.drop('Latitude', axis=1)
    data_to_model = data_to_model.drop('Longitude', axis=1)
    data_to_model = data_to_model.drop('Link', axis=1)

    datasize = data_to_model.shape[0]

    data_to_model.insert(2, 'Huff_predict', list(i for i in range(datasize)))

    data_to_model.insert(2, 'Name', list('name' for i in range(datasize)))

    for i in range(datasize):
        data_to_model['Huff_predict'][i] = data_huff_pred(i, lamb, alp)

    for i in range(datasize):
        data_to_model['Name'][i] = name

    data_pred = model.predict(data_to_model)

    beststores = dict()
    predictions = []
    for i, item in enumerate(data_pred):
        if item in beststores.keys():
            predictions.append(item)
            if dataset['Square'][i] > dataset['Square'][beststores[item]]:
                beststores[item] = i
        else:
            beststores[item] = i
    predictions.sort(reverse=True)
    i = 1
    results = []
    last = 0
    for elem in predictions:
        if i == 6:
            break
        if elem != last:
            last = elem
            result = {}
            result['Топ'] = i
            result['Предполагаемая посещаемость'] = round(data_pred[beststores[elem]])
            result['Стоимость одного посетителя'] = dataset['Price'][beststores[elem]] / (
                        30 * round(data_pred[beststores[elem]]))
            result['Стоимость аренды'] = dataset['Price'][beststores[elem]]
            result['Площадь'] = dataset['Square'][beststores[elem]]
            result['Адрес'] = dataset['Address'][beststores[elem]]
            result['Широта'] = dataset['Latitude'][beststores[elem]]
            result['Долгота'] = dataset['Longitude'][beststores[elem]]
            result['Ссылка на объявление'] = dataset['Link'][beststores[elem]]
            results.append(result)
            i += 1
    return results[:5]
